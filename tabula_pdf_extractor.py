#!/usr/bin/python3


import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
import codecs
import os
import sys
import xlsxwriter
import tabula

'''
Extraction example for PDF tables shown in raw.pdf example
https://pypi.org/project/tabula-py/
'''

if len(sys.argv) == 1:

    print ("")
    print ("")
    print ("PDF filename not passed as PDF. Please pass PDF filename as first argument OR Parameter")
    print ("")
    print ("")
    sys.exit()


pdf = sys.argv[1]
#extract filename without extension
pdfn = pdf.split(".")[0]

os.system('mkdir input')
os.system('cp ' + pdf + ' input/')
os.system('chmod -R 777 input')
#Split pdf using pdftk cmd
os.system('cd input && pdftk ' + pdf + ' burst')
os.remove(os.path.join('input',pdf))
os.system('rm -Rf input/*.txt')
os.system('rm -Rf input/*.jpg')
os.system('chmod -R 777 `pwd`/input')


#Converting the splitted PDF into list
csvlist = sorted(os.listdir('input'))


#Create Excel Workbook
workbook = xlsxwriter.Workbook('tables.xlsx')
worksheet = workbook.add_worksheet()
workbook.close()
os.system('chmod -R 777 tables.xlsx')

#Load excel Workbook using openpyxl
book = load_workbook('tables.xlsx')
writer = ExcelWriter('tables.xlsx', engine='openpyxl') 	
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

#Iterating through the csv list
for i in csvlist:

    #converting csv into DataFrame
    df = tabula.read_pdf(os.path.join('input',i))

    #converting a DataFrame to excel
    df.to_excel(writer,sheet_name="table_" + str(i),index=False,header=True)
	
writer.save()
